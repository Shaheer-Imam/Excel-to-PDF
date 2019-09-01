import openpyxl
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
import os
from pyPDF2 import PdfFileReader, PdfFileMerger

pdfmetrics.registerFont(TTFont('Arial','Arial.ttf'))
wb = openpyxl.load_workbook('F:\\Python\\py to pdf\\Data.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

#print(sheet.cell(2,2).value)

page_width = 2156
page_height = 3050
start = 200
start_2 = 700
spread = 80
categories = ["Roll Number: ","Name: ","Gender: ","Age: ","Board: ","City: ","Email: "]

university = "NED University Of Engineering And Technology"

def create_data():
    for i in range (2,14):
         std_id = sheet.cell(row=i,column=1).value
         std_name = sheet.cell(row=i,column=2).value
         std_gender = sheet.cell(row=i,column=3).value
         std_age = sheet.cell(row=i,column=4).value
         std_board = sheet.cell(row=i,column=5).value
         std_city = sheet.cell(row=i,column=6).value
         std_email = sheet.cell(row=i,column=8).value

         data=[std_id,std_name,std_gender,std_age,std_board,std_city,std_email]

         c=canvas.Canvas(str(std_name)+'.pdf')
         c.setPageSize((page_width,page_height))
         c.setFont('Arial',80)
         text_width = stringWidth(university,'Arial',80)
         c.drawString((page_width-text_width)/2,2900,university)

         y=2500

         for x in range(0,7):
             c.setFont('Arial',40)
             c.drawString(start,y,categories[x])
             c.drawString(start_2,y,str(data[x])) 
             y-=spread


         c.save()
         
def merge_data():
    files_dir = 'F:\\Python\\py to pdf'
    pdf_files = [f for f in os.listdir(files_dir) if f.endswith('.pdf')]
    merger = PdfFileMerger()
    for filename in pdf_files:
        merger.append(PdfFileReader(os.path.join(files_dir,filename,'rb')))
    merger.write(os.path.join(files_dir,'merged_data.pdf'))
         
create_data()
merge_data()
